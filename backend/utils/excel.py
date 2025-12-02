"""
Excel 导入导出工具
"""
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from typing import List, Dict
from io import BytesIO
from datetime import date, datetime


def export_to_excel(data: List[Dict], headers: Dict[str, str], filename: str = "export.xlsx") -> BytesIO:
    """
    导出数据到 Excel
    
    Args:
        data: 数据列表，每个元素为字典
        headers: 表头映射，key为字段名，value为显示名称
        filename: 文件名
    
    Returns:
        BytesIO: Excel 文件二进制流
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "数据"
    
    # 设置表头样式
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # 写入表头
    col_keys = list(headers.keys())
    for col_idx, key in enumerate(col_keys, start=1):
        cell = ws.cell(row=1, column=col_idx, value=headers[key])
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # 写入数据
    for row_idx, item in enumerate(data, start=2):
        for col_idx, key in enumerate(col_keys, start=1):
            value = item.get(key, "")
            
            # 处理日期类型
            if isinstance(value, (date, datetime)):
                value = value.strftime("%Y-%m-%d")
            
            # 处理 None 值
            if value is None:
                value = ""
            
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # 自动调整列宽
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # 保存到内存
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output


def import_from_excel(file: BytesIO, required_headers: List[str]) -> List[Dict]:
    """
    从 Excel 导入数据
    
    Args:
        file: Excel 文件二进制流
        required_headers: 必需的表头列表
    
    Returns:
        List[Dict]: 导入的数据列表
    """
    wb = load_workbook(file)
    ws = wb.active
    
    # 读取表头
    headers = []
    for cell in ws[1]:
        headers.append(cell.value)
    
    # 验证表头
    for required in required_headers:
        if required not in headers:
            raise ValueError(f"缺少必需的列：{required}")
    
    # 读取数据
    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(cell is None for cell in row):  # 跳过空行
            continue
        
        row_data = {}
        for idx, value in enumerate(row):
            if idx < len(headers):
                row_data[headers[idx]] = value
        
        data.append(row_data)
    
    return data


def export_papers_to_excel(papers: List) -> BytesIO:
    """
    导出论文数据到 Excel
    """
    data = []
    for paper in papers:
        data.append({
            "id": paper.id,
            "title": paper.title,
            "authors": paper.authors,
            "journal": paper.journal,
            "publication_date": paper.publication_date,
            "doi": paper.doi,
            "jcr_zone": paper.jcr_zone,
            "cas_zone": paper.cas_zone,
            "impact_factor": paper.impact_factor,
        })
    
    headers = {
        "id": "ID",
        "title": "论文标题",
        "authors": "作者",
        "journal": "期刊",
        "publication_date": "发表日期",
        "doi": "DOI",
        "jcr_zone": "JCR分区",
        "cas_zone": "中科院分区",
        "impact_factor": "影响因子",
    }
    
    return export_to_excel(data, headers, "论文数据.xlsx")


def export_projects_to_excel(projects: List) -> BytesIO:
    """
    导出项目数据到 Excel
    """
    data = []
    for project in projects:
        data.append({
            "id": project.id,
            "project_name": project.project_name,
            "pi_name": project.pi_name,
            "project_type": project.project_type,
            "source": project.source,
            "budget_total": project.budget_total,
            "start_date": project.start_date,
            "end_date": project.end_date,
            "status": project.status.value if hasattr(project.status, 'value') else project.status,
        })
    
    headers = {
        "id": "ID",
        "project_name": "项目名称",
        "pi_name": "负责人",
        "project_type": "项目类型",
        "source": "项目来源",
        "budget_total": "总预算",
        "start_date": "开始日期",
        "end_date": "结束日期",
        "status": "状态",
    }
    
    return export_to_excel(data, headers, "项目数据.xlsx")


def export_achievements_to_excel(achievements: List) -> BytesIO:
    """
    导出成果数据到 Excel
    """
    data = []
    for achievement in achievements:
        data.append({
            "id": achievement.id,
            "achievement_type": achievement.achievement_type.value if hasattr(achievement.achievement_type, 'value') else achievement.achievement_type,
            "title": achievement.title,
            "owner": achievement.owner,
            "members": achievement.members,
            "completion_date": achievement.completion_date,
            "certificate_no": achievement.certificate_no,
        })
    
    headers = {
        "id": "ID",
        "achievement_type": "成果类型",
        "title": "成果名称",
        "owner": "所有人",
        "members": "参与人员",
        "completion_date": "完成日期",
        "certificate_no": "证书编号",
    }
    
    return export_to_excel(data, headers, "成果数据.xlsx")
